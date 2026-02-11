# -*- coding: utf-8 -*-
"""
把 data/ 下的 CSV 转成带 K 线图 + MA20 的 Excel，同一文件里既有数据表又有图。
用 Excel / WPS 打开 xlsx 即可，不依赖 HTML。
"""

import os
from openpyxl import Workbook
from openpyxl.chart import StockChart, LineChart, Reference
from openpyxl.chart.axis import ChartLines

from config import OUT_EXCEL, OUT_EXCEL_ALT, SYMBOL_LIST as SYMBOLS, csv_path


def load_csv(symbol: str, name: str) -> list:
    path = csv_path(symbol, name)
    if not os.path.exists(path):
        return []
    rows = []
    with open(path, "r", encoding="utf-8-sig") as f:
        next(f)
        for line in f:
            line = line.strip()
            if not line:
                continue
            parts = line.split(",")
            if len(parts) < 6:
                continue
            try:
                date = parts[0].strip()
                open_ = float(parts[1])
                high = float(parts[2])
                low = float(parts[3])
                close = float(parts[4])
                vol = int(float(parts[5]))
            except (ValueError, TypeError):
                continue
            rows.append((date, open_, high, low, close, vol))
    return rows


def add_ma20(rows: list) -> list:
    """每行增加 MA20 列；不足 20 行为空。"""
    out = []
    closes = [r[4] for r in rows]
    for i, r in enumerate(rows):
        if i < 19:
            ma = None
        else:
            ma = round(sum(closes[i - 19 : i + 1]) / 20, 2)
        out.append((*r, ma))
    return out


def main():
    wb = Workbook()
    wb.remove(wb.active)
    for symbol, name in SYMBOLS:
        data = add_ma20(load_csv(symbol, name))
        if len(data) < 2:
            continue
        ws = wb.create_sheet(title=name[:6], index=len(wb.worksheets))
        # 表头：日期, 开盘(元/吨), 最高, 最低, 收盘, 成交量, MA20
        headers = ("日期", "开盘(元/吨)", "最高(元/吨)", "最低(元/吨)", "收盘(元/吨)", "成交量(手)", "MA20")
        for c, h in enumerate(headers, 1):
            ws.cell(row=1, column=c, value=h)
        for r, row in enumerate(data, 2):
            for c, val in enumerate(row, 1):
                ws.cell(row=r, column=c, value=val)
        n = len(data) + 1
        row_start, row_end = 2, n
        # K 线（高-低-收）A=日期 B=开 C=高 D=低 E=收 F=量 G=MA20
        c1 = StockChart()
        labels = Reference(ws, min_col=1, min_row=row_start, max_row=row_end)
        data_ref = Reference(ws, min_col=3, max_col=5, min_row=1, max_row=row_end)
        c1.add_data(data_ref, titles_from_data=True)
        c1.set_categories(labels)
        c1.hiLowLines = ChartLines()
        c1.title = f"{name} 日K（高-低-收）+ MA20"
        c1.width = 18
        c1.height = 12
        ws.add_chart(c1, "H2")
        # MA20 折线图（放在 K 线图下方，同一列）
        c2 = LineChart()
        c2.title = "MA20"
        ma_ref = Reference(ws, min_col=7, min_row=1, max_row=row_end)
        c2.add_data(ma_ref, titles_from_data=True)
        c2.set_categories(labels)
        c2.width = 18
        c2.height = 6
        ws.add_chart(c2, "H16")
    if not wb.worksheets:
        print("未找到任何 CSV 数据")
        return
    base = os.path.dirname(__file__)
    out_path = os.path.join(base, OUT_EXCEL)
    try:
        wb.save(out_path)
    except PermissionError:
        out_path = os.path.join(base, OUT_EXCEL_ALT)
        wb.save(out_path)
        print(f"原文件可能被占用，已另存为: {OUT_EXCEL_ALT}")
    print(f"已生成: {out_path}")
    print("用 Excel 或 WPS 打开，每张表左侧为完整数据表（含 MA20 列），右侧为 K 线图 + MA20 线。")


if __name__ == "__main__":
    main()
